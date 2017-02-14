# !/usr/bin/python
# coding:utf-8
import datetime
import getopt
import sys
# from openpyxl import load_workbook
try:
	from openpyxl import load_workbook
except ImportError:
	sys.stderr.write("Please install openpyxl with 'pip install openpyxl'.\n")
	sys.exit(2)

"""批注导出工具：
1. 支持从多个工作表导出批注
2. 导出批注根据各工作表生成相应批注汇总工作表
3. 导出批注另存为新的工作簿，避免对已有工作薄修改

执行：
python export_comments.py excel_name

如需修改批注导出格式，请更新'headline' 和'dataline' 列表，注意保持匹配关系，如：
[ '序号', '批注所在位置', '批注生成时间', '评审人员', '批注内容', '问题级别', '备注']
[index, comment_data.parent.coordinate, datetime.datetime.now(), comment_data.author, comment_data.text, '', '']
"""

VERSION = '0.0.1'

# 批注标题栏
headline = ['序号', '批注所在位置', '批注生成时间', '评审人员', '批注内容', '问题级别', '备注']

def usage():
	print("Usage: " + __file__ + """ excel_file
Please provide excel_file, 请提供Excel文件
""")
	sys.exit(2)

def read_comments(workbook, worksheet_list):
	"""
	读取批注
	:param workbook: obj, 工作表对象
	:return: list, 批注列表
	"""
	all_comments = {}

	for worksheet_name in worksheet_list:
		worksheet = workbook[worksheet_name]
		rows_data = tuple(worksheet.rows)

		comment_list = []
		for row_line in rows_data:
			for cell in row_line:
				if cell.comment:
					comment_list.append(cell.comment)

		if comment_list:
			all_comments[worksheet_name] = comment_list

	return all_comments

def write_comments(comment_list, worksheet_to_write):
	"""
	批注内容写入新表单
	:param comment_list: list 批注列表
	:param worksheet: obj 表单对象
	:return:
	"""

	max_col = len(headline)
	max_row = len(comment_list) + 1    # 标题行处理，在数据行数基础上+1
	# 初始化/创建单元格区域
	create_cell = worksheet_to_write.iter_rows(min_row=1, max_col=max_col, max_row=max_row)

	for index, rows in enumerate(create_cell):
		i = 0
		for cell in rows:
			if index == 0:
				# 填充数据，首行标题
				cell.value = headline[i]
				i += 1
			else:
				comment_data = comment_list[index-1]    # 标题行处理， 在数据行数基础上-1
				data_line = [index, comment_data.parent.coordinate, datetime.datetime.now(), comment_data.author,
							 comment_data.text, '', '']
				cell.value = data_line[i]
				i += 1

def main():
	try:
		(opts, args) = getopt.getopt(sys.argv[1:], '')
	except getopt.GetoptError:
		usage()

	if len(args) != 1:
		usage()

	input_file_name = args[0]

	workbook = load_workbook(filename=input_file_name)
	worksheet_list = workbook.sheetnames
	all_comments = read_comments(workbook, worksheet_list)

	for name, comments in all_comments.items():
		worksheet_to_write = workbook.create_sheet(name+'_comments')
		write_comments(comments, worksheet_to_write)

	name, ext_format = input_file_name.split('.')
	workbook.save(name + '_comments.' + ext_format)

main()